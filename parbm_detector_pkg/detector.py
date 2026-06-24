"""
Core PARBM Detector functionality with publication-quality visualizations.
"""

import requests
import json
import re
import csv
from typing import Dict, List, Tuple, Optional, Union
from urllib.parse import quote
from collections import defaultdict
from dataclasses import dataclass, asdict

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import matplotlib.gridspec as gridspec
    from matplotlib.patches import FancyBboxPatch, FancyArrowPatch
    from matplotlib.colors import LinearSegmentedColormap, Normalize
    import matplotlib.cm as cm
    import numpy as np
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

from .input_handler import InputHandler


# ── Palettes & style ──────────────────────────────────────────────────────────

PARBM_COLOR   = "#E63946"   # crimson  – PAR-binding domains
DOMAIN_COLOR  = "#457B9D"   # steel-blue – other domains
HIGH_CONF_COLOR = "#F4A261" # amber    – high-confidence sites
LOW_CONF_COLOR  = "#A8DADC" # teal     – lower-confidence sites
BG_COLOR      = "#F8F9FA"
PANEL_BG      = "#FFFFFF"
TEXT_DARK     = "#1D3557"
GRID_COLOR    = "#DEE2E6"

MOTIF_PALETTE = {
    "SxxE motif":  "#E63946",
    "SxxD motif":  "#F4A261",
    "TxxE motif":  "#2A9D8F",
    "ExxE motif":  "#457B9D",
    "Acidic region": "#9B5DE5",
}

# ── Biology constants ─────────────────────────────────────────────────────────

PARBM_DOMAINS = {
    # WWE domain (PAR-binding)
    "IPR004170": "WWE domain",
    "IPR018123": "WWE domain, subgroup",
    # Macro/MacroD domain (PAR-binding)
    "IPR002589": "Macro domain",
    "IPR015398": "MacroD-type domain",
    # BRCT domain (PAR-binding, e.g. XRCC1, PARP1)
    "IPR002126": "BRCT domain",
    "IPR036895": "BRCT domain superfamily",
    # PBZ / PAR-binding zinc finger
    "IPR034642": "PAR-binding zinc finger (PBZ)",
    # RRM / RNA recognition motif (some PAR readers)
    "IPR000504": "RNA recognition motif (RRM)",
    # Zinc finger CCCH (some PAR-associated)
    "IPR001650": "Zinc finger CCCH",
    # PARP catalytic / regulatory
    "IPR012317": "PARP catalytic domain",
    "IPR024486": "PARP regulatory domain",
}

PARYATION_MOTIFS = {
    "SxxE": (r"S.{2}E", "SxxE motif"),
    "SxxD": (r"S.{2}D", "SxxD motif"),
    "TxxE": (r"T.{2}E", "TxxE motif"),
    "ExxE": (r"E.{2}E", "ExxE motif"),
    "acidic_rich": (r"[DE]{3,}", "Acidic region"),
}

HYDROPHOBICITY = {
    "A": 1.8,  "R": -4.5, "N": -3.5, "D": -3.5, "C": 2.5,
    "Q": -3.5, "E": -3.5, "G": -0.4, "H": -3.2, "I": 4.5,
    "L": 3.8,  "K": -3.9, "M": 1.9,  "F": 2.8,  "P": -1.6,
    "S": -0.8, "T": -0.7, "W": -0.9, "Y": -1.3, "V": 4.2,
}

AA_CHARGE = {
    "D": -1, "E": -1, "K": 1, "R": 1, "H": 0.1,
}

AA_GROUPS = {
    "hydrophobic": set("AILMFWV"),
    "polar":       set("STNQCY"),
    "charged_pos": set("KRH"),
    "charged_neg": set("DE"),
    "special":     set("GP"),
}


# ── Data classes ──────────────────────────────────────────────────────────────

@dataclass
class PARylationSite:
    position: int
    residue: str
    motif_type: str
    score: float


@dataclass
class DomainHit:
    name: str
    interpro_id: str
    start: int
    end: int
    e_value: float = 0.0
    is_parbm: bool = False


# ── Predictor ─────────────────────────────────────────────────────────────────

class PARylationSitePredictor:
    """Predicts PARylation sites from sequence features."""

    @staticmethod
    def _window(sequence: str, pos: int, w: int) -> str:
        return sequence[max(0, pos - w): min(len(sequence), pos + w + 1)]

    @staticmethod
    def calculate_local_charge(sequence: str, position: int, window: int = 5) -> float:
        local = PARylationSitePredictor._window(sequence, position, window)
        charge = sum(AA_CHARGE.get(aa, 0) for aa in local)
        return charge / len(local) if local else 0

    @staticmethod
    def calculate_local_hydrophobicity(sequence: str, position: int, window: int = 3) -> float:
        local = PARylationSitePredictor._window(sequence, position, window)
        h = sum(HYDROPHOBICITY.get(aa, 0) for aa in local)
        return h / len(local) if local else 0

    @staticmethod
    def predict_sites(sequence: str) -> List[PARylationSite]:
        motif_positions: Dict[int, List[Tuple[str, float]]] = defaultdict(list)

        for motif_name, (pattern, motif_type) in PARYATION_MOTIFS.items():
            for match in re.finditer(pattern, sequence, re.IGNORECASE):
                for pos in range(match.start(), match.end()):
                    if sequence[pos] in "STED":
                        motif_positions[pos].append((motif_type, 0.8))

        sites = []
        for pos, motifs in motif_positions.items():
            residue = sequence[pos]
            motif_type, base_score = motifs[0]

            charge   = PARylationSitePredictor.calculate_local_charge(sequence, pos)
            hydro    = PARylationSitePredictor.calculate_local_hydrophobicity(sequence, pos)

            charge_penalty = max(0, charge * 0.1)
            hydro_bonus    = 0.05 if hydro < 0 else 0
            multi_bonus    = min(0.1, (len(motifs) - 1) * 0.05)

            score = max(0, min(1, base_score - charge_penalty + hydro_bonus + multi_bonus))
            sites.append(PARylationSite(pos + 1, residue, motif_type, score))

        sites.sort(key=lambda x: x.score, reverse=True)
        return sites


# ── Sliding-window helper ─────────────────────────────────────────────────────

def _sliding(sequence: str, func, window: int = 9) -> List[float]:
    n = len(sequence)
    out = []
    for i in range(n):
        lo, hi = max(0, i - window // 2), min(n, i + window // 2 + 1)
        chunk = sequence[lo:hi]
        out.append(func(chunk))
    return out


# ── Main detector ─────────────────────────────────────────────────────────────

class PARBMDetector:
    """Multi-input PAR-binding motif detector with rich visualizations."""

    def __init__(self):
        self.uniprot_api  = "https://rest.uniprot.org/uniprotkb"
        self.interpro_api = "https://www.ebi.ac.uk/interpro/api/entry/interpro/protein/UniProt"
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": "PARBMDetector/3.0"})
        self.predictor     = PARylationSitePredictor()
        self.input_handler = InputHandler()

    # ── Fetching ─────────────────────────────────────────────────────────────

    def fetch_uniprot_sequence(self, uniprot_id: str) -> Optional[Dict]:
        try:
            url    = f"{self.uniprot_api}/{quote(uniprot_id)}"
            params = {"format": "json", "fields": "accession,id,protein_name,organism_name,sequence,gene_names"}
            r = self.session.get(url, params=params, timeout=15)
            r.raise_for_status()
            return r.json()
        except requests.exceptions.RequestException as e:
            print(f"  ✗ UniProt fetch error: {e}")
            return None

    def fetch_interpro_domains(self, uniprot_id: str) -> Optional[Dict]:
        try:
            url = f"{self.interpro_api}/{quote(uniprot_id)}"
            r   = self.session.get(url, timeout=15)
            r.raise_for_status()
            return r.json()
        except requests.exceptions.RequestException as e:
            print(f"  ✗ InterPro fetch error: {e}")
            return None

    def extract_domains(self, interpro_data: Optional[Dict]) -> List[DomainHit]:
        if not interpro_data or "results" not in interpro_data:
            return []
        domains = []
        for result in interpro_data.get("results", []):
            meta   = result.get("metadata", {})
            ipr_id = meta.get("accession", "")
            name   = meta.get("name", "Unknown")
            for protein in result.get("proteins", []):
                for loc in protein.get("entry_protein_locations", []):
                    for frag in loc.get("fragments", [{}]):
                        domains.append(DomainHit(
                            name=name,
                            interpro_id=ipr_id,
                            start=frag.get("start", 0),
                            end=frag.get("end", 0),
                            e_value=loc.get("e_value", 0.0),
                            is_parbm=ipr_id in PARBM_DOMAINS,
                        ))
        return domains

    # ── Analysis ─────────────────────────────────────────────────────────────

    def analyze_from_uniprot_id(self, uniprot_id: str) -> Dict:
        print(f"\n  Analyzing UniProt ID: {uniprot_id}")
        data = self.fetch_uniprot_sequence(uniprot_id)
        if not data:
            return {"error": f"Could not fetch data for {uniprot_id}"}

        sequence     = data.get("sequence", {}).get("value", "")
        protein_name = (data.get("proteinDescription", {})
                            .get("recommendedName", {})
                            .get("fullName", {})
                            .get("value", "Unknown"))
        org_data     = data.get("organism", {})
        sci_name     = org_data.get("scientificName", "Unknown")
        common_name  = org_data.get("commonName", "")
        organism     = f"{sci_name} ({common_name})" if common_name else sci_name
        gene_name    = (data.get("genes", [{}])[0]
                            .get("geneName", {})
                            .get("value", "N/A")
                        if data.get("genes") else "N/A")

        domains       = self.extract_domains(self.fetch_interpro_domains(uniprot_id))
        parbm_domains = [d for d in domains if d.is_parbm]
        sites         = self.predictor.predict_sites(sequence)
        hc_sites      = [s for s in sites if s.score >= 0.7]

        return self._build_result(
            uniprot_id, protein_name, gene_name, organism, sequence,
            "uniprot_id", domains, parbm_domains, sites, hc_sites,
        )

    def analyze_from_sequence(self, sequence: str,
                               protein_name: str = "Unknown",
                               gene_name: str = "N/A",
                               organism: str = "Unknown") -> Dict:
        print(f"\n  Analyzing sequence: {protein_name} ({len(sequence)} aa)")
        valid_aa = set("ACDEFGHIKLMNPQRSTVWY")
        sequence = "".join(c for c in sequence.upper().replace("-", "").replace("*", "")
                           if c in valid_aa)
        sites    = self.predictor.predict_sites(sequence)
        hc_sites = [s for s in sites if s.score >= 0.7]
        return self._build_result(
            None, protein_name, gene_name, organism, sequence,
            "sequence", [], [], sites, hc_sites,
        )

    def _build_result(self, uid, name, gene, org, seq, itype,
                      domains, parbm, sites, hc_sites) -> Dict:
        top_site = sites[0] if sites else None
        parbm_label = (
            ", ".join(PARBM_DOMAINS.get(d.interpro_id, d.name) for d in parbm)
            if parbm else "None detected (no WWE/MacroD/BRCT/ZnF-CCCH)"
        )
        return {
            "uniprot_id": uid,
            "protein_name": name,
            "gene_name": gene,
            "organism": org,
            "sequence": seq,
            "sequence_length": len(seq),
            "input_type": itype,
            "all_domains": domains,
            "parbm_domains": parbm,
            "paryation_sites": sites,
            "high_confidence_sites": hc_sites,
            "summary": {
                "total_domains": len(domains),
                "parbm_domains_count": len(parbm),
                "parbm_domains_label": parbm_label,
                "total_paryation_sites": len(sites),
                "high_confidence_sites_count": len(hc_sites),
                "top_site": (
                    f"{top_site.residue}{top_site.position} via "
                    f"{top_site.motif_type}, score {top_site.score:.2f}"
                ) if top_site else "N/A",
            },
        }

    def analyze(self, input_data: Union[str, List]) -> Union[Dict, List[Dict]]:
        if isinstance(input_data, list):
            results = []
            for item in input_data:
                r = self.analyze(item)
                results.extend(r if isinstance(r, list) else [r])
            return results

        s = str(input_data).strip()
        t = self.input_handler.detect_input_type(s)

        if t == "uniprot_id":
            return self.analyze_from_uniprot_id(s)
        elif t == "sequence":
            return self.analyze_from_sequence(s)
        elif t in ("fasta", "fasta_file"):
            entries = (self.input_handler.parse_fasta(s)
                       if t == "fasta"
                       else self.input_handler.read_fasta_file(s))
            return [self.analyze_from_sequence(seq, header) for header, seq in entries]
        else:
            return {"error": f"Unrecognised input: {s[:50]}"}

    def batch_analyze(self, inputs: List) -> List[Dict]:
        print(f"\n{'='*70}\nBATCH PROCESSING: {len(inputs)} input(s)\n{'='*70}")
        results = []
        for item in inputs:
            r = self.analyze(item)
            results.extend(r if isinstance(r, list) else [r])
        return results

    # ── Print ─────────────────────────────────────────────────────────────────

    def print_analysis(self, result: Dict):
        if "error" in result:
            print(f"  ✗ {result['error']}")
            return

        s = result["summary"]
        sep = "─" * 56

        print(f"\n  {sep}")
        print(f"  {'PARBM DETECTOR — ANALYSIS REPORT':^54}")
        print(f"  {sep}")

        # ── Protein identity ──────────────────────────────────────
        print(f"  {'Protein':<32} {result['protein_name']}")
        print(f"  {'Gene':<32} {result.get('gene_name', 'N/A')}")
        print(f"  {'UniProt ID':<32} {result.get('uniprot_id') or 'N/A'}")
        print(f"  {'Organism':<32} {result['organism']}")
        print(f"  {'Amino acids':<32} {result['sequence_length']}")
        print(f"  {'Input type':<32} {result['input_type']}")

        # ── PARylation summary ────────────────────────────────────
        print(f"\n  {sep}")
        print(f"  {'PAR-binding domains':<32} {s['parbm_domains_label']}")
        hc_count = s['high_confidence_sites_count']
        hc_label = f"{hc_count} site{'s' if hc_count != 1 else ''}"
        print(f"  {'High-conf sites (≥0.70)':<32} {hc_label}")
        print(f"  {'All predicted sites':<32} {s['total_paryation_sites']}")
        print(f"  {'Top site':<32} {s['top_site']}")

        # ── Site detail ───────────────────────────────────────────
        hc = result["high_confidence_sites"]
        if hc:
            print(f"\n  High-confidence PARylation sites:")
            print(f"  {'Residue':<10} {'Motif':<18} {'Score':>6}")
            print(f"  {'·'*8:<10} {'·'*16:<18} {'·'*5:>6}")
            for site in hc[:10]:
                print(f"  {site.residue+str(site.position):<10} {site.motif_type:<18} {site.score:>6.2f}")
            if len(hc) > 10:
                print(f"  … and {len(hc) - 10} more")

        # ── PAR-binding domain detail ─────────────────────────────
        if result["parbm_domains"]:
            print(f"\n  PAR-binding domain locations:")
            for d in result["parbm_domains"]:
                label = PARBM_DOMAINS.get(d.interpro_id, d.name)
                print(f"  · {label:<30} [{d.start}–{d.end}]")

        print(f"  {sep}\n")

    # ── Export ────────────────────────────────────────────────────────────────

    def export_results(self, results: List[Dict], output_file: str = "parbm_results.json"):
        export = []
        for r in results:
            row = dict(r)
            if "error" not in row:
                row["all_domains"]          = [asdict(d) for d in row.get("all_domains", [])]
                row["parbm_domains"]        = [asdict(d) for d in row.get("parbm_domains", [])]
                row["paryation_sites"]      = [asdict(s) for s in row.get("paryation_sites", [])]
                row["high_confidence_sites"]= [asdict(s) for s in row.get("high_confidence_sites", [])]
            export.append(row)
        with open(output_file, "w") as f:
            json.dump(export, f, indent=2)
        print(f"\n✓ Results exported → {output_file}")

    def generate_csv_report(self, results: List[Dict], output_file: str = "parbm_sites_report.csv"):
        with open(output_file, "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["Protein_ID", "Protein_Name", "Input_Type",
                         "Position", "Residue", "Motif_Type",
                         "Confidence_Score", "Sequence_Length"])
            for r in results:
                if "error" not in r:
                    for s in r["paryation_sites"]:
                        w.writerow([
                            r.get("uniprot_id", "N/A"),
                            r["protein_name"][:50],
                            r["input_type"],
                            s.position, s.residue, s.motif_type,
                            f"{s.score:.3f}", r["sequence_length"],
                        ])
        print(f"✓ CSV report exported → {output_file}")

    # ══════════════════════════════════════════════════════════════════════════
    # VISUALIZATION  –  publication-quality multi-panel dashboard
    # ══════════════════════════════════════════════════════════════════════════

    def visualize_protein(self, result: Dict, output_file: str = None,
                          show: bool = False) -> Optional[str]:
        """
        Generate a comprehensive 6-panel figure:
          Panel A – domain architecture map
          Panel B – PARylation site confidence landscape
          Panel C – sliding-window hydrophobicity profile
          Panel D – local charge profile with site markers
          Panel E – motif-type breakdown (horizontal bar)
          Panel F – summary statistics & sequence composition
        """
        if not MATPLOTLIB_AVAILABLE:
            print("  ✗ matplotlib / numpy not available.")
            return None
        if "error" in result:
            print(f"  ✗ {result['error']}")
            return None

        # ── unpack ────────────────────────────────────────────────────────────
        name    = result["protein_name"]
        seq     = result["sequence"]
        seqlen  = result["sequence_length"]
        domains = result["all_domains"]
        parbm   = result["parbm_domains"]
        sites   = result["paryation_sites"]
        hc      = result["high_confidence_sites"]

        if not seq:
            print("  ✗ No sequence available for visualisation.")
            return None

        positions = np.array([s.position for s in sites]) if sites else np.array([])
        scores    = np.array([s.score    for s in sites]) if sites else np.array([])

        # sliding-window profiles
        hydro_win  = 11
        charge_win = 15
        hydro_profile  = _sliding(seq, lambda w: sum(HYDROPHOBICITY.get(a, 0) for a in w) / len(w), hydro_win)
        charge_profile = _sliding(seq, lambda w: sum(AA_CHARGE.get(a, 0) for a in w) / len(w), charge_win)
        x_pos = np.arange(1, seqlen + 1)

        # ── build shared panel data (reused by individual saves too) ────────────
        uid_str      = result.get("uniprot_id") or "N/A"
        gene_str     = result.get("gene_name") or "N/A"
        org_str      = result.get("organism", "Unknown")
        s            = result["summary"]
        hc_count     = s["high_confidence_sites_count"]
        top_site_str = s.get("top_site", "N/A")

        def _style_ax(ax):
            ax.set_facecolor(PANEL_BG)
            for sp in ax.spines.values():
                sp.set_color(GRID_COLOR)
                sp.set_linewidth(0.8)

        # ── figure layout ─────────────────────────────────────────────────────
        # Row 0  : title bar (tall enough so text never bleeds into Row 1)
        # Row 1  : info table (identity + summary side-by-side)
        # Rows 2–5: science panels A–F
        fig = plt.figure(figsize=(18, 25), facecolor=BG_COLOR)
        fig.patch.set_facecolor(BG_COLOR)

        gs_outer = gridspec.GridSpec(
            6, 1,
            figure=fig,
            height_ratios=[0.18, 1.1, 1, 1, 1, 1],
            hspace=0.50,
            left=0.06, right=0.97,
            top=0.99, bottom=0.32,   # bottom 30% reserved for panels G & H
        )

        ax_title = fig.add_subplot(gs_outer[0])  # narrow title stripe
        ax_info  = fig.add_subplot(gs_outer[1])  # info table row
        ax_info.set_clip_on(True)

        # 4-row × 2-col grid for science panels
        gs_plots = gridspec.GridSpecFromSubplotSpec(
            4, 2,
            subplot_spec=gs_outer[2:],
            hspace=0.52,
            wspace=0.32,
        )
        ax_arch   = fig.add_subplot(gs_plots[0, :])
        ax_conf   = fig.add_subplot(gs_plots[1, :])
        ax_hydro  = fig.add_subplot(gs_plots[2, 0])
        ax_charge = fig.add_subplot(gs_plots[2, 1])
        ax_motif  = fig.add_subplot(gs_plots[3, 0])
        ax_stats  = fig.add_subplot(gs_plots[3, 1])

        for ax in [ax_arch, ax_conf, ax_hydro, ax_charge, ax_motif, ax_stats]:
            _style_ax(ax)

        # ── TITLE STRIPE ──────────────────────────────────────────────────────
        for sp in ax_title.spines.values():
            sp.set_visible(False)
        ax_title.set_xticks([]); ax_title.set_yticks([])
        ax_title.set_facecolor("#1D3557")
        ax_title.text(
            0.5, 0.5,
            "PARBM Detector v3.0  ·  PARylation & PAR-binding Motif Analysis",
            transform=ax_title.transAxes,
            fontsize=13, fontweight="bold", color="white",
            ha="center", va="center",
        )

        # ── INFO TABLE ROW ─────────────────────────────────────────────────────
        for sp in ax_info.spines.values():
            sp.set_color(GRID_COLOR); sp.set_linewidth(0.8)
        ax_info.set_facecolor("#EBF5FB")
        ax_info.set_xticks([]); ax_info.set_yticks([])

        # left column – identity (6 rows of key/value)
        id_rows = [
            ("Protein",    name),
            ("Gene",       gene_str),
            ("UniProt ID", uid_str),
            ("Organism",   org_str),
            ("Length",     f"{seqlen} aa"),
            ("Input type", result["input_type"]),
        ]
        row_h  = 1 / len(id_rows)
        for i, (k, v) in enumerate(id_rows):
            y = 1 - (i + 0.5) * row_h
            bg = "#D6EAF8" if i % 2 == 0 else "#EBF5FB"
            ax_info.axhspan(1 - (i + 1) * row_h, 1 - i * row_h,
                            xmin=0, xmax=0.5, color=bg, zorder=0)
            ax_info.text(0.01, y, k, transform=ax_info.transAxes,
                         fontsize=10, fontweight="bold", color=TEXT_DARK,
                         va="center", ha="left", family="monospace")
            ax_info.text(0.26, y, v, transform=ax_info.transAxes,
                         fontsize=10, color=TEXT_DARK,
                         va="center", ha="left", family="monospace",
                         clip_on=True)

        # right column – PARylation summary
        hc_label  = f"{hc_count} site{'s' if hc_count != 1 else ''}"
        sum_rows  = [
            ("PAR-binding domains",    s["parbm_domains_label"]),
            ("High-conf sites (≥0.70)", hc_label),
            ("All predicted sites",    str(s["total_paryation_sites"])),
            ("Top site",               top_site_str),
        ]
        row_h2 = 1 / len(sum_rows)
        for i, (k, v) in enumerate(sum_rows):
            y = 1 - (i + 0.5) * row_h2
            bg = "#FDEBD0" if i % 2 == 0 else "#FEF9E7"
            ax_info.axhspan(1 - (i + 1) * row_h2, 1 - i * row_h2,
                            xmin=0.5, xmax=1.0, color=bg, zorder=0)
            ax_info.text(0.51, y, k, transform=ax_info.transAxes,
                         fontsize=10, fontweight="bold", color=TEXT_DARK,
                         va="center", ha="left", family="monospace")
            ax_info.text(0.74, y, v, transform=ax_info.transAxes,
                         fontsize=10, color="#E63946" if i == 3 else TEXT_DARK,
                         va="center", ha="left", family="monospace",
                         clip_on=True)

        # vertical divider
        ax_info.axvline(0.5, color=GRID_COLOR, lw=1.2)

        # panel letters – placed just above each science panel inside figure coords,
        # computed after layout so they sit clear of the info row
        _panel_label = dict(fontsize=13, fontweight="bold", color=TEXT_DARK,
                            transform=fig.transFigure, va="top")

        # ══════════════════════════════════════════════════════════════════════
        # Panel A – Domain architecture
        # ══════════════════════════════════════════════════════════════════════
        ax = ax_arch

        # ── greedy track assignment so overlapping domains never share a row ──
        def assign_tracks(domain_list):
            """Return list of (domain, track_index) with no x-overlap per track."""
            tracks = []   # each entry = rightmost end used on that track
            assignment = []
            for dom in domain_list:
                placed = False
                for ti, track_end in enumerate(tracks):
                    if dom.start > track_end + 2:
                        tracks[ti] = dom.end
                        assignment.append(ti)
                        placed = True
                        break
                if not placed:
                    tracks.append(dom.end)
                    assignment.append(len(tracks) - 1)
            return assignment

        track_idx  = assign_tracks(domains) if domains else []
        n_tracks   = max(track_idx) + 1 if track_idx else 1

        # y-layout:  backbone at centre, tracks fanned above and below
        # even track indices → above backbone, odd → below
        BACKBONE_Y  = 0.50
        TRACK_STEP  = 0.18   # vertical gap between track centres
        BOX_H       = 0.13   # height of each domain box

        def track_y(ti):
            level = (ti // 2) + 1
            return BACKBONE_Y + level * TRACK_STEP if ti % 2 == 0 \
                   else BACKBONE_Y - level * TRACK_STEP

        # compute y-limits dynamically
        if track_idx:
            all_ys = [track_y(ti) for ti in track_idx]
            y_lo = min(all_ys) - BOX_H - 0.08
            y_hi = max(all_ys) + BOX_H + 0.12
        else:
            y_lo, y_hi = BACKBONE_Y - 0.25, BACKBONE_Y + 0.25

        # backbone line
        ax.plot([1, seqlen], [BACKBONE_Y, BACKBONE_Y],
                color="#868E96", lw=3, solid_capstyle="round", zorder=2)

        # N / C termini
        ax.text(1,      BACKBONE_Y, "N", ha="center", va="center",
                fontsize=9, fontweight="bold", color="#495057", zorder=5,
                bbox=dict(boxstyle="circle,pad=0.15", fc="white", ec="#868E96", lw=1))
        ax.text(seqlen, BACKBONE_Y, "C", ha="center", va="center",
                fontsize=9, fontweight="bold", color="#495057", zorder=5,
                bbox=dict(boxstyle="circle,pad=0.15", fc="white", ec="#868E96", lw=1))

        # domain boxes + connector lines + external labels
        cmap_conf = LinearSegmentedColormap.from_list(
            "conf", ["#FFCF77", "#E63946"], N=256
        )
        for domain, ti in zip(domains, track_idx):
            is_p  = domain in parbm
            col   = PARBM_COLOR if is_p else DOMAIN_COLOR
            dy    = track_y(ti)
            dw    = max(domain.end - domain.start, 3)
            mid   = (domain.start + domain.end) / 2

            # thin connector from backbone to domain box
            ax.plot([mid, mid], [BACKBONE_Y, dy],
                    color=col, lw=1.0, ls="--", alpha=0.45, zorder=2)

            # domain box
            ax.add_patch(FancyBboxPatch(
                (domain.start, dy - BOX_H / 2), dw, BOX_H,
                boxstyle="round,pad=0.008",
                facecolor=col, edgecolor="white", linewidth=1.2,
                alpha=0.90, zorder=4,
            ))

            # label: prefer inside box; fall back to above/below if box too narrow
            label = PARBM_DOMAINS.get(domain.interpro_id, domain.name)
            # use short label inside, full label as annotation outside
            short = label if len(label) <= 18 else label[:16] + "…"
            if dw > seqlen * 0.08:   # wide enough for inside label
                ax.text(mid, dy, short,
                        ha="center", va="center",
                        fontsize=8, fontweight="bold", color="white",
                        zorder=5, clip_on=True)
            else:
                # label outside, with a small pointer
                label_y = dy + BOX_H / 2 + 0.04 if dy >= BACKBONE_Y \
                          else dy - BOX_H / 2 - 0.04
                va_out  = "bottom" if dy >= BACKBONE_Y else "top"
                ax.annotate(
                    short, xy=(mid, dy),
                    xytext=(mid, label_y),
                    ha="center", va=va_out,
                    fontsize=7.5, fontweight="bold", color=col,
                    arrowprops=dict(arrowstyle="-", color=col, lw=0.7),
                    zorder=6,
                )

        # site markers on backbone
        for s in sites:
            c  = cmap_conf(s.score)
            ms = 9 if s.score >= 0.7 else 5
            ax.plot(s.position, BACKBONE_Y, "v",
                    color=c, markersize=ms,
                    markeredgecolor="white", markeredgewidth=0.5,
                    alpha=0.95 if s.score >= 0.7 else 0.55, zorder=6)

        # colorbar
        sm = plt.cm.ScalarMappable(cmap=cmap_conf, norm=Normalize(vmin=0, vmax=1))
        sm.set_array([])
        cb = fig.colorbar(sm, ax=ax, orientation="vertical",
                          fraction=0.015, pad=0.01, aspect=15)
        cb.set_label("Site confidence", fontsize=8, color=TEXT_DARK)
        cb.ax.tick_params(labelsize=7, colors=TEXT_DARK)

        ax.set_xlim(-seqlen * 0.02, seqlen * 1.03)
        ax.set_ylim(y_lo, y_hi)
        ax.set_xlabel("Residue position", fontsize=9, color=TEXT_DARK)
        ax.set_yticks([])
        ax.set_title("A  ·  Domain Architecture & PARylation Sites", fontsize=10,
                     fontweight="bold", color=TEXT_DARK, pad=6)
        ax.grid(axis="x", color=GRID_COLOR, linewidth=0.5, alpha=0.5)
        ax.tick_params(colors=TEXT_DARK, labelsize=8)

        # legend
        handles = [
            mpatches.Patch(facecolor=PARBM_COLOR,  label="PAR-binding domain"),
            mpatches.Patch(facecolor="#457B9D", label="Other domain"),
            plt.Line2D([0], [0], marker="v", lw=0,
                       markerfacecolor="#E63946", markeredgecolor="white",
                       markersize=8, label="High-conf site (≥0.7)"),
            plt.Line2D([0], [0], marker="v", lw=0,
                       markerfacecolor="#FFCF77", markeredgecolor="white",
                       markersize=6, label="Lower-conf site"),
        ]
        ax.legend(handles=handles, loc="upper right", fontsize=7.5,
                  framealpha=0.85, edgecolor=GRID_COLOR)

        # ══════════════════════════════════════════════════════════════════════
        # Panel B – Confidence landscape
        # ══════════════════════════════════════════════════════════════════════
        ax = ax_conf
        if len(positions) > 0:
            # density-style score landscape using a gaussian kernel
            score_arr = np.zeros(seqlen)
            sigma = max(seqlen / 120, 3)
            for pos, sc in zip(positions, scores):
                idx = int(pos) - 1
                kernel_range = range(max(0, idx - int(4 * sigma)),
                                     min(seqlen, idx + int(4 * sigma) + 1))
                for j in kernel_range:
                    score_arr[j] = max(score_arr[j],
                                       sc * np.exp(-0.5 * ((j - idx) / sigma) ** 2))

            # fill under curve
            ax.fill_between(x_pos, score_arr, alpha=0.25, color=PARBM_COLOR)
            ax.plot(x_pos, score_arr, lw=1.2, color=PARBM_COLOR, alpha=0.7)

            # individual site stems
            for pos, sc in zip(positions, scores):
                col = PARBM_COLOR if sc >= 0.7 else HIGH_CONF_COLOR
                ax.vlines(pos, 0, sc, colors=col, lw=1.0, alpha=0.7)
                ax.plot(pos, sc, "o", color=col, markersize=5 if sc >= 0.7 else 3,
                        markeredgecolor="white", markeredgewidth=0.4, zorder=5)

            ax.axhline(0.7, color=PARBM_COLOR, lw=1.2, ls="--", alpha=0.8,
                       label="High-confidence threshold (0.70)")
            ax.axhline(0.5, color="#ADB5BD", lw=0.8, ls=":", alpha=0.7,
                       label="Moderate threshold (0.50)")

            # shade domain regions
            for domain in domains:
                col = PARBM_COLOR if domain in parbm else DOMAIN_COLOR
                ax.axvspan(domain.start, domain.end, alpha=0.06, color=col)

            ax.set_ylim(0, 1.08)
            ax.legend(fontsize=8, loc="upper right", framealpha=0.85, edgecolor=GRID_COLOR)
        else:
            ax.text(0.5, 0.5, "No PARylation sites predicted",
                    ha="center", va="center", fontsize=12, color="#6C757D",
                    transform=ax.transAxes)

        ax.set_xlim(1, seqlen)
        ax.set_xlabel("Residue position", fontsize=9, color=TEXT_DARK)
        ax.set_ylabel("Confidence score", fontsize=9, color=TEXT_DARK)
        ax.set_title("B  ·  PARylation Site Confidence Landscape", fontsize=10,
                     fontweight="bold", color=TEXT_DARK, pad=6)
        ax.grid(color=GRID_COLOR, linewidth=0.5, alpha=0.6)
        ax.tick_params(colors=TEXT_DARK, labelsize=8)

        # ══════════════════════════════════════════════════════════════════════
        # Panel C – Hydrophobicity profile
        # ══════════════════════════════════════════════════════════════════════
        ax = ax_hydro
        hp = np.array(hydro_profile)
        ax.fill_between(x_pos, hp, where=hp >= 0, alpha=0.35,
                        color="#F4A261", label="Hydrophobic")
        ax.fill_between(x_pos, hp, where=hp < 0,  alpha=0.35,
                        color="#457B9D", label="Hydrophilic")
        ax.plot(x_pos, hp, lw=0.9, color=TEXT_DARK, alpha=0.6)
        ax.axhline(0, color="#6C757D", lw=0.8, ls="--")

        for s in hc:
            ax.axvline(s.position, color=PARBM_COLOR, lw=0.7, alpha=0.4)

        ax.set_xlim(1, seqlen)
        ax.set_xlabel("Residue position", fontsize=8, color=TEXT_DARK)
        ax.set_ylabel(f"Hydrophobicity\n(window={hydro_win})", fontsize=8, color=TEXT_DARK)
        ax.set_title("C  ·  Sliding-window Hydrophobicity", fontsize=9,
                     fontweight="bold", color=TEXT_DARK, pad=4)
        ax.legend(fontsize=7, framealpha=0.85, edgecolor=GRID_COLOR)
        ax.grid(color=GRID_COLOR, linewidth=0.4, alpha=0.5)
        ax.tick_params(colors=TEXT_DARK, labelsize=7)

        # ══════════════════════════════════════════════════════════════════════
        # Panel D – Charge profile
        # ══════════════════════════════════════════════════════════════════════
        ax = ax_charge
        cp = np.array(charge_profile)
        ax.fill_between(x_pos, cp, where=cp >= 0, alpha=0.35,
                        color="#2A9D8F", label="Positive")
        ax.fill_between(x_pos, cp, where=cp < 0,  alpha=0.35,
                        color="#E63946", label="Negative")
        ax.plot(x_pos, cp, lw=0.9, color=TEXT_DARK, alpha=0.6)
        ax.axhline(0, color="#6C757D", lw=0.8, ls="--")

        for s in hc:
            ax.axvline(s.position, color="#9B5DE5", lw=0.7, alpha=0.35)

        ax.set_xlim(1, seqlen)
        ax.set_xlabel("Residue position", fontsize=8, color=TEXT_DARK)
        ax.set_ylabel(f"Net charge\n(window={charge_win})", fontsize=8, color=TEXT_DARK)
        ax.set_title("D  ·  Local Charge Profile", fontsize=9,
                     fontweight="bold", color=TEXT_DARK, pad=4)
        ax.legend(fontsize=7, framealpha=0.85, edgecolor=GRID_COLOR)
        ax.grid(color=GRID_COLOR, linewidth=0.4, alpha=0.5)
        ax.tick_params(colors=TEXT_DARK, labelsize=7)

        # ══════════════════════════════════════════════════════════════════════
        # Panel E – Motif-type breakdown
        # ══════════════════════════════════════════════════════════════════════
        ax = ax_motif
        motif_counts: Dict[str, int] = defaultdict(int)
        motif_scores: Dict[str, List[float]] = defaultdict(list)
        for s in sites:
            motif_counts[s.motif_type] += 1
            motif_scores[s.motif_type].append(s.score)

        if motif_counts:
            labels  = list(motif_counts.keys())
            counts  = [motif_counts[l] for l in labels]
            colors  = [MOTIF_PALETTE.get(l, "#6C757D") for l in labels]
            avg_sc  = [np.mean(motif_scores[l]) for l in labels]

            # sort by count descending
            order   = sorted(range(len(counts)), key=lambda i: counts[i], reverse=True)
            labels  = [labels[i]  for i in order]
            counts  = [counts[i]  for i in order]
            colors  = [colors[i]  for i in order]
            avg_sc  = [avg_sc[i]  for i in order]

            y_pos   = np.arange(len(labels))
            bars    = ax.barh(y_pos, counts, color=colors, edgecolor="white",
                              linewidth=0.8, height=0.65, alpha=0.88)

            # annotate bars with average score
            for bar, sc in zip(bars, avg_sc):
                w = bar.get_width()
                ax.text(w + 0.15, bar.get_y() + bar.get_height() / 2,
                        f"avg {sc:.2f}", va="center", ha="left",
                        fontsize=7.5, color=TEXT_DARK)

            ax.set_yticks(y_pos)
            ax.set_yticklabels(labels, fontsize=8, color=TEXT_DARK)
            ax.set_xlabel("Site count", fontsize=8, color=TEXT_DARK)
            ax.set_xlim(0, max(counts) * 1.35)
        else:
            ax.text(0.5, 0.5, "No motifs detected",
                    ha="center", va="center", fontsize=10, color="#6C757D",
                    transform=ax.transAxes)

        ax.set_title("E  ·  Motif-type Breakdown", fontsize=9,
                     fontweight="bold", color=TEXT_DARK, pad=4)
        ax.grid(axis="x", color=GRID_COLOR, linewidth=0.4, alpha=0.5)
        ax.tick_params(colors=TEXT_DARK, labelsize=7)

        # ══════════════════════════════════════════════════════════════════════
        # Panel F – Summary stats + residue composition pie
        # ══════════════════════════════════════════════════════════════════════
        ax = ax_stats

        # compute AA group composition
        group_counts = {g: sum(1 for aa in seq if aa in aas)
                        for g, aas in AA_GROUPS.items()}
        total_aa = sum(group_counts.values())
        group_fracs = {g: v / total_aa * 100 for g, v in group_counts.items()}

        pie_labels = ["Hydrophobic", "Polar", "Pos. charged", "Neg. charged", "Special (G/P)"]
        pie_values = [group_fracs["hydrophobic"], group_fracs["polar"],
                      group_fracs["charged_pos"], group_fracs["charged_neg"],
                      group_fracs["special"]]
        pie_colors = ["#F4A261", "#2A9D8F", "#457B9D", "#E63946", "#9B5DE5"]

        wedges, texts, autotexts = ax.pie(
            pie_values,
            labels=pie_labels,
            colors=pie_colors,
            autopct="%1.0f%%",
            startangle=90,
            pctdistance=0.75,
            wedgeprops=dict(edgecolor="white", linewidth=1.2),
            textprops=dict(fontsize=7.5, color=TEXT_DARK),
        )
        for at in autotexts:
            at.set_fontsize(7)
            at.set_color("white")
            at.set_fontweight("bold")

        # stats text box – site statistics (identity is already in the header panel)
        stats_lines = [
            f"PAR-binding domains : {len(parbm)}",
            f"Other domains       : {len(domains) - len(parbm)}",
            f"─" * 26,
            f"All predicted sites : {len(sites)}",
            f"High-conf (≥ 0.70)  : {len(hc)}",
        ]
        if sites:
            stats_lines += [
                f"Top score           : {max(s.score for s in sites):.3f}",
                f"Mean score          : {np.mean([s.score for s in sites]):.3f}",
            ]
        top_site_str = result["summary"].get("top_site", "N/A")
        if top_site_str and top_site_str != "N/A":
            # wrap long top-site string
            stats_lines.append(f"Top site            : {top_site_str}")

        stats_text = "\n".join(stats_lines)
        ax.text(
            0.5, -0.18, stats_text,
            transform=ax.transAxes,
            fontsize=8.5, color=TEXT_DARK,
            va="top", ha="center",
            bbox=dict(boxstyle="round,pad=0.6", facecolor=BG_COLOR,
                      edgecolor=GRID_COLOR, linewidth=0.8),
            family="monospace",
        )

        ax.set_title("F  ·  Residue Composition & Summary", fontsize=9,
                     fontweight="bold", color=TEXT_DARK, pad=4)

        # ══════════════════════════════════════════════════════════════════════
        # Panel G – Disorder prediction overlay
        # ══════════════════════════════════════════════════════════════════════
        # Add two more rows to the outer grid
        enrich     = result.get("enrichment", {})
        dis_data   = enrich.get("disorder", {})
        dis_scores = dis_data.get("scores", [])
        dis_method = dis_data.get("method", "")

        # Panels G & H sit below the 6-row subplot grid.
        # gs_outer occupies top=0.99, bottom=0.03.
        # We allocated 6 rows in gs_outer with height_ratios [0.18, 1.1, 1,1,1,1].
        # The two new panels occupy the bottom band we deliberately left free.
        ax_disorder = fig.add_axes([0.06, 0.215, 0.86, 0.092])
        ax_residue  = fig.add_axes([0.06, 0.108, 0.86, 0.092])

        for axx in [ax_disorder, ax_residue]:
            _style_ax(axx)

        # G – disorder
        ax = ax_disorder
        if dis_scores and len(dis_scores) == seqlen:
            xd = np.arange(1, seqlen + 1)
            yd = np.array(dis_scores)
            ax.fill_between(xd, yd, where=yd >= 0.5, alpha=0.35, color="#9B5DE5",
                            label="Disordered (>0.5)")
            ax.fill_between(xd, yd, where=yd < 0.5, alpha=0.20, color="#ADB5BD",
                            label="Ordered")
            ax.plot(xd, yd, lw=0.9, color="#9B5DE5", alpha=0.7)
            ax.axhline(0.5, color="#6C757D", lw=0.8, ls="--")
            for site in hc:
                if site.position <= seqlen:
                    ax.axvline(site.position, color=PARBM_COLOR, lw=0.6, alpha=0.35)
            ax.legend(fontsize=7, framealpha=0.85, edgecolor=GRID_COLOR,
                      loc="upper right")
            ax.set_ylim(0, 1.08)
            mean_d = dis_data.get("mean_disorder", 0)
            frac_d = dis_data.get("disordered_fraction", 0)
            ax.text(0.01, 0.92,
                    f"Mean={mean_d:.2f}  Disordered={frac_d*100:.0f}%  ({dis_method})",
                    transform=ax.transAxes, fontsize=7.5, color="#6C757D")
        else:
            ax.text(0.5, 0.5, "Disorder data not available — run enrich_result()",
                    ha="center", va="center", fontsize=9, color="#6C757D",
                    transform=ax.transAxes)
        ax.set_xlim(1, seqlen)
        ax.set_xlabel("Residue position", fontsize=8, color=TEXT_DARK)
        ax.set_ylabel("Disorder score", fontsize=8, color=TEXT_DARK)
        ax.set_title("G  ·  Intrinsic Disorder Profile (IUPred2A)", fontsize=9,
                     fontweight="bold", color=TEXT_DARK, pad=4)
        ax.grid(color=GRID_COLOR, linewidth=0.4, alpha=0.5)
        ax.tick_params(colors=TEXT_DARK, labelsize=7)

        # ══════════════════════════════════════════════════════════════════════
        # Panel H – Residue-type breakdown (S/T/E/D/Y/K/R)
        # ══════════════════════════════════════════════════════════════════════
        ax = ax_residue
        from collections import Counter
        aa_order  = ["S", "T", "E", "D", "Y", "K", "R"]
        aa_colors = [PARBM_COLOR, HIGH_CONF_COLOR, "#2A9D8F", "#457B9D",
                     "#9B5DE5", "#20B2AA", "#FF8C00"]
        all_counts = Counter(site.residue for site in sites)
        hc_counts  = Counter(site.residue for site in hc)
        x_pos_h    = np.arange(len(aa_order))
        bar_w      = 0.35

        bars_all = ax.bar(x_pos_h - bar_w / 2,
                          [all_counts.get(aa, 0) for aa in aa_order],
                          width=bar_w, label="All sites",
                          color=aa_colors, alpha=0.45,
                          edgecolor="white", linewidth=0.8)
        bars_hc  = ax.bar(x_pos_h + bar_w / 2,
                          [hc_counts.get(aa, 0) for aa in aa_order],
                          width=bar_w, label="High-conf (≥0.70)",
                          color=aa_colors,
                          edgecolor="white", linewidth=0.8)

        # annotate bars
        for bar in list(bars_all) + list(bars_hc):
            h = bar.get_height()
            if h > 0:
                ax.text(bar.get_x() + bar.get_width() / 2, h + 0.15,
                        str(int(h)), ha="center", va="bottom",
                        fontsize=7, color=TEXT_DARK)

        ax.set_xticks(x_pos_h)
        ax.set_xticklabels(aa_order, fontsize=9, fontweight="bold")
        ax.set_ylabel("Site count", fontsize=8, color=TEXT_DARK)
        ax.set_title("H  ·  PARylation Sites by Residue Type", fontsize=9,
                     fontweight="bold", color=TEXT_DARK, pad=4)
        ax.legend(fontsize=7.5, framealpha=0.85, edgecolor=GRID_COLOR)
        ax.grid(axis="y", color=GRID_COLOR, linewidth=0.4, alpha=0.5)
        ax.tick_params(colors=TEXT_DARK, labelsize=8)

        # ── save individual panels ────────────────────────────────────────────
        if output_file is None:
            safe = re.sub(r"[^\w\-]", "_", name[:30])
            output_file = f"parbm_{safe}.png"

        import os
        out_dir   = os.path.dirname(output_file) or "."
        base_stem = re.sub(r"\.png$", "", os.path.basename(output_file))

        panel_axes = [
            (ax_arch,     "A_domain_architecture"),
            (ax_conf,     "B_confidence_landscape"),
            (ax_hydro,    "C_hydrophobicity"),
            (ax_charge,   "D_charge_profile"),
            (ax_motif,    "E_motif_breakdown"),
            (ax_stats,    "F_residue_composition"),
            (ax_disorder, "G_disorder_profile"),
            (ax_residue,  "H_residue_type_breakdown"),
        ]

        panels_dir = os.path.join(out_dir, "panels")
        os.makedirs(panels_dir, exist_ok=True)

        renderer = fig.canvas.get_renderer()
        for panel_ax, panel_name in panel_axes:
            extent = panel_ax.get_tightbbox(renderer).transformed(
                fig.dpi_scale_trans.inverted()
            )
            p_png = os.path.join(panels_dir, f"{base_stem}_{panel_name}.png")
            p_pdf = os.path.join(panels_dir, f"{base_stem}_{panel_name}.pdf")
            p_svg = os.path.join(panels_dir, f"{base_stem}_{panel_name}.svg")
            fig.savefig(p_png, dpi=200, bbox_inches=extent, facecolor=BG_COLOR)
            fig.savefig(p_pdf, bbox_inches=extent, facecolor=BG_COLOR)
            fig.savefig(p_svg, bbox_inches=extent, facecolor=BG_COLOR)
            print(f"  ✓ Panel {panel_name.split('_')[0]} → {os.path.basename(p_png)}")

        # ── save full composite ───────────────────────────────────────────────
        plt.savefig(output_file, dpi=200, bbox_inches="tight", facecolor=BG_COLOR)
        print(f"  ✓ Composite PNG → {output_file}")

        pdf_file = re.sub(r"\.png$", ".pdf", output_file)
        plt.savefig(pdf_file, bbox_inches="tight", facecolor=BG_COLOR)
        print(f"  ✓ Composite PDF → {pdf_file}")

        svg_file = re.sub(r"\.png$", ".svg", output_file)
        plt.savefig(svg_file, bbox_inches="tight", facecolor=BG_COLOR)
        print(f"  ✓ Composite SVG → {svg_file}")

        if show:
            plt.show()
        plt.close(fig)
        return output_file, pdf_file

    # ── Batch figure ──────────────────────────────────────────────────────────

    def visualize_comparison(self, results: List[Dict], output_file: str = "parbm_comparison.png"):
        """
        Side-by-side confidence-score landscape for multiple proteins.
        """
        if not MATPLOTLIB_AVAILABLE:
            print("  ✗ matplotlib not available.")
            return
        valid = [r for r in results if "error" not in r and r.get("sequence")]
        if not valid:
            print("  ✗ No valid results to compare.")
            return

        n    = len(valid)
        fig, axes = plt.subplots(n, 1, figsize=(16, 3.5 * n), facecolor=BG_COLOR)
        if n == 1:
            axes = [axes]
        fig.suptitle("Comparative PARylation Confidence Profiles",
                     fontsize=13, fontweight="bold", color=TEXT_DARK, y=1.01)

        for ax, r in zip(axes, valid):
            ax.set_facecolor(PANEL_BG)
            seqlen   = r["sequence_length"]
            sites    = r["paryation_sites"]
            x_pos    = np.arange(1, seqlen + 1)

            if sites:
                score_arr = np.zeros(seqlen)
                sigma     = max(seqlen / 120, 3)
                for s in sites:
                    idx = s.position - 1
                    for j in range(max(0, idx - int(4*sigma)),
                                   min(seqlen, idx + int(4*sigma) + 1)):
                        score_arr[j] = max(score_arr[j],
                                          s.score * np.exp(-0.5 * ((j - idx) / sigma) ** 2))
                ax.fill_between(x_pos, score_arr, alpha=0.3, color=PARBM_COLOR)
                ax.plot(x_pos, score_arr, lw=1.0, color=PARBM_COLOR)

            for d in r.get("parbm_domains", []):
                ax.axvspan(d.start, d.end, alpha=0.12, color=PARBM_COLOR)
            ax.axhline(0.7, color=PARBM_COLOR, lw=0.8, ls="--", alpha=0.6)

            uid  = f" ({r['uniprot_id']})" if r.get("uniprot_id") else ""
            ax.set_title(f"{r['protein_name']}{uid}  [{seqlen} aa]",
                         fontsize=9, fontweight="bold", color=TEXT_DARK)
            ax.set_xlim(1, seqlen)
            ax.set_ylim(0, 1.05)
            ax.set_ylabel("Confidence", fontsize=8, color=TEXT_DARK)
            ax.grid(color=GRID_COLOR, linewidth=0.4, alpha=0.5)
            ax.tick_params(colors=TEXT_DARK, labelsize=7)
            for sp in ax.spines.values():
                sp.set_color(GRID_COLOR)

        axes[-1].set_xlabel("Residue position", fontsize=9, color=TEXT_DARK)

        plt.tight_layout()
        plt.savefig(output_file, dpi=180, bbox_inches="tight", facecolor=BG_COLOR)
        print(f"  ✓ Comparison figure saved → {output_file}")
        plt.close(fig)

    # ══════════════════════════════════════════════════════════════════════════
    # Excel report
    # ══════════════════════════════════════════════════════════════════════════

    def export_excel(self, result: Dict, output_file: str) -> str:
        """Write a formatted multi-sheet Excel workbook for one protein result."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side, numbers
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("  ✗ openpyxl not installed. Run: pip install openpyxl")
            return None

        if "error" in result:
            print(f"  ✗ {result['error']}")
            return None

        wb = Workbook()

        # ── colour palette ────────────────────────────────────────────────────
        C_HEADER_BG  = "1D3557"   # dark navy
        C_HEADER_FG  = "FFFFFF"
        C_PARBM_BG   = "E63946"   # crimson
        C_PARBM_FG   = "FFFFFF"
        C_ALT_BG     = "EBF5FB"   # light blue-grey
        C_SUBHEAD_BG = "A8DADC"   # teal
        C_SUBHEAD_FG = "1D3557"
        C_BORDER     = "ADB5BD"

        thin = Side(style="thin", color=C_BORDER)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def hdr_fill(hex_col):
            return PatternFill("solid", fgColor=hex_col)

        def hdr_font(hex_col="FFFFFF", bold=True, size=11):
            return Font(color=hex_col, bold=bold, size=size, name="Calibri")

        def center(wrap=False):
            return Alignment(horizontal="center", vertical="center",
                             wrap_text=wrap)

        def left(wrap=False):
            return Alignment(horizontal="left", vertical="center",
                             wrap_text=wrap)

        def set_col_width(ws, col, width):
            ws.column_dimensions[get_column_letter(col)].width = width

        # ══════════════════════════════════════════════════════════════════════
        # Sheet 1 – Summary
        # ══════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.sheet_view.showGridLines = False

        s   = result["summary"]
        uid = result.get("uniprot_id") or "N/A"
        hc  = result["high_confidence_sites"]

        # Title banner (merged A1:D1)
        ws1.merge_cells("A1:D1")
        c = ws1["A1"]
        c.value = "PARBM Detector v3.0  —  Analysis Report"
        c.font      = Font(color=C_HEADER_FG, bold=True, size=15, name="Calibri")
        c.fill      = hdr_fill(C_HEADER_BG)
        c.alignment = center()
        ws1.row_dimensions[1].height = 30

        # Sub-title
        ws1.merge_cells("A2:D2")
        c = ws1["A2"]
        c.value = "PAR-binding Motif & PARylation Site Analysis"
        c.font      = Font(color="A8DADC", bold=False, size=11, name="Calibri")
        c.fill      = hdr_fill(C_HEADER_BG)
        c.alignment = center()
        ws1.row_dimensions[2].height = 20

        # blank row
        ws1.row_dimensions[3].height = 8

        # ── Identity block ────────────────────────────────────────────────────
        def section_header(ws, row, label, cols="A:D"):
            ws.merge_cells(f"A{row}:D{row}")
            c = ws[f"A{row}"]
            c.value     = label
            c.font      = hdr_font(C_SUBHEAD_FG, bold=True, size=10)
            c.fill      = hdr_fill(C_SUBHEAD_BG)
            c.alignment = left()
            ws.row_dimensions[row].height = 18

        def kv_row(ws, row, key, value, alt=False):
            bg = C_ALT_BG if alt else "FFFFFF"
            for col, val in enumerate([key, "", str(value), ""], start=1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill      = hdr_fill(bg)
                c.alignment = left(wrap=True)
                c.border    = border
                c.font      = Font(name="Calibri", size=10,
                                   bold=(col == 1), color="1D3557")
            ws.merge_cells(f"B{row}:B{row}")  # keep col B empty
            ws.merge_cells(f"C{row}:D{row}")  # value spans C–D
            ws.row_dimensions[row].height = 16

        section_header(ws1, 4, "  Protein Identity")
        kv_row(ws1, 5,  "Protein",    result["protein_name"])
        kv_row(ws1, 6,  "Gene",       result.get("gene_name", "N/A"), alt=True)
        kv_row(ws1, 7,  "UniProt ID", uid)
        kv_row(ws1, 8,  "Organism",   result["organism"], alt=True)
        kv_row(ws1, 9,  "Length (aa)",result["sequence_length"])
        kv_row(ws1, 10, "Input type", result["input_type"], alt=True)

        ws1.row_dimensions[11].height = 8

        section_header(ws1, 12, "  PARylation Summary")
        kv_row(ws1, 13, "PAR-binding domains",     s["parbm_domains_label"])
        kv_row(ws1, 14, "High-confidence sites (≥ 0.70)",
               s["high_confidence_sites_count"], alt=True)
        kv_row(ws1, 15, "All predicted sites",     s["total_paryation_sites"])
        kv_row(ws1, 16, "Top site",                s.get("top_site", "N/A"), alt=True)

        # PAR-binding domain detail
        if result["parbm_domains"]:
            ws1.row_dimensions[17].height = 8
            section_header(ws1, 18, "  PAR-binding Domain Locations")
            for i, dom in enumerate(result["parbm_domains"], start=19):
                label = PARBM_DOMAINS.get(dom.interpro_id, dom.name)
                kv_row(ws1, i, label, f"[{dom.start}–{dom.end}]", alt=(i % 2 == 0))

        # column widths
        for col, w in [(1, 34), (2, 4), (3, 36), (4, 4)]:
            set_col_width(ws1, col, w)

        # ══════════════════════════════════════════════════════════════════════
        # Sheet 2 – PARylation Sites
        # ══════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("PARylation Sites")
        ws2.sheet_view.showGridLines = False

        # title
        ws2.merge_cells("A1:F1")
        c = ws2["A1"]
        c.value = (f"PARylation Sites  —  {result['protein_name']} "
                   f"({uid})  |  {result['sequence_length']} aa")
        c.font      = Font(color=C_HEADER_FG, bold=True, size=13, name="Calibri")
        c.fill      = hdr_fill(C_HEADER_BG)
        c.alignment = center()
        ws2.row_dimensions[1].height = 26

        headers = ["#", "Position", "Residue", "Motif Type", "Confidence Score", "Category"]
        for col, h in enumerate(headers, start=1):
            c = ws2.cell(row=2, column=col, value=h)
            c.font      = hdr_font(size=10)
            c.fill      = hdr_fill("2C3E6B")
            c.alignment = center()
            c.border    = border
        ws2.row_dimensions[2].height = 18

        sites_all = result["paryation_sites"]
        for i, site in enumerate(sites_all, start=1):
            row = i + 2
            category = "High confidence" if site.score >= 0.7 else "Moderate"
            row_bg = "FFF3CD" if site.score >= 0.9 else (
                     C_ALT_BG if site.score >= 0.7 else "FFFFFF")
            vals = [i, site.position, site.residue,
                    site.motif_type, round(site.score, 3), category]
            for col, val in enumerate(vals, start=1):
                c = ws2.cell(row=row, column=col, value=val)
                c.fill      = hdr_fill(row_bg)
                c.alignment = center()
                c.border    = border
                c.font      = Font(name="Calibri", size=10,
                                   bold=(site.score >= 0.9),
                                   color=("E63946" if site.score >= 0.9 else "1D3557"))
            ws2.row_dimensions[row].height = 15

        for col, w in [(1,5),(2,11),(3,11),(4,20),(5,18),(6,18)]:
            set_col_width(ws2, col, w)

        # ══════════════════════════════════════════════════════════════════════
        # Sheet 3 – Domains
        # ══════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("Domains")
        ws3.sheet_view.showGridLines = False

        ws3.merge_cells("A1:F1")
        c = ws3["A1"]
        c.value = f"Domain Architecture  —  {result['protein_name']} ({uid})"
        c.font      = Font(color=C_HEADER_FG, bold=True, size=13, name="Calibri")
        c.fill      = hdr_fill(C_HEADER_BG)
        c.alignment = center()
        ws3.row_dimensions[1].height = 26

        d_headers = ["#", "Domain Name", "InterPro ID", "Start", "End", "PAR-binding?"]
        for col, h in enumerate(d_headers, start=1):
            c = ws3.cell(row=2, column=col, value=h)
            c.font      = hdr_font(size=10)
            c.fill      = hdr_fill("2C3E6B")
            c.alignment = center()
            c.border    = border
        ws3.row_dimensions[2].height = 18

        for i, dom in enumerate(result["all_domains"], start=1):
            row    = i + 2
            is_par = dom in result["parbm_domains"]
            row_bg = "FFD6D6" if is_par else ("FFFFFF" if i % 2 else C_ALT_BG)
            vals   = [i, dom.name, dom.interpro_id, dom.start, dom.end,
                      "YES ✓" if is_par else "No"]
            for col, val in enumerate(vals, start=1):
                c = ws3.cell(row=row, column=col, value=val)
                c.fill      = hdr_fill(row_bg)
                c.alignment = center()
                c.border    = border
                c.font      = Font(name="Calibri", size=10,
                                   bold=is_par,
                                   color=(C_PARBM_BG if is_par else "1D3557"))
            ws3.row_dimensions[row].height = 15

        if not result["all_domains"]:
            ws3.cell(row=3, column=1,
                     value="No domain annotations retrieved (sequence input or API unavailable)")

        for col, w in [(1,5),(2,38),(3,16),(4,9),(5,9),(6,15)]:
            set_col_width(ws3, col, w)

        # ══════════════════════════════════════════════════════════════════════
        # Sheet 4 – Sequence
        # ══════════════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet("Sequence")
        ws4.sheet_view.showGridLines = False

        ws4.merge_cells("A1:B1")
        c = ws4["A1"]
        c.value     = f"Protein Sequence  —  {result['protein_name']}"
        c.font      = Font(color=C_HEADER_FG, bold=True, size=13, name="Calibri")
        c.fill      = hdr_fill(C_HEADER_BG)
        c.alignment = center()
        ws4.row_dimensions[1].height = 26

        seq = result.get("sequence", "")
        hc_positions = {s.position for s in result["high_confidence_sites"]}
        chunk = 60
        for i, start in enumerate(range(0, len(seq), chunk)):
            row      = i + 2
            segment  = seq[start: start + chunk]
            label    = f"{start + 1}–{min(start + chunk, len(seq))}"
            c_label  = ws4.cell(row=row, column=1, value=label)
            c_label.font      = Font(name="Courier New", size=9,
                                     bold=True, color="6C757D")
            c_label.alignment = left()
            c_seq = ws4.cell(row=row, column=2, value=segment)
            c_seq.font      = Font(name="Courier New", size=9, color="1D3557")
            c_seq.alignment = left()
            ws4.row_dimensions[row].height = 14

        set_col_width(ws4, 1, 12)
        set_col_width(ws4, 2, 72)

        wb.save(output_file)
        print(f"  ✓ Excel saved → {output_file}")
        return output_file

    # ══════════════════════════════════════════════════════════════════════════
    # Master export — one folder per protein
    # ══════════════════════════════════════════════════════════════════════════

    def export_to_folder(self, result: Dict, base_dir: str = ".") -> str:
        """
        Create a dedicated output folder for one protein and write:
          • <ID>_<Gene>_analysis.png   – publication PNG (200 dpi)
          • <ID>_<Gene>_analysis.pdf   – vector PDF
          • <ID>_<Gene>_report.xlsx    – formatted multi-sheet workbook
        Returns the folder path.
        """
        import os

        if "error" in result:
            print(f"  ✗ {result['error']}")
            return None

        uid  = result.get("uniprot_id") or "SEQ"
        gene = result.get("gene_name") or result["protein_name"][:12]
        safe_gene = re.sub(r"[^\w]", "_", gene)
        folder_name = f"{uid}_{safe_gene}"
        folder = os.path.join(base_dir, folder_name)
        os.makedirs(folder, exist_ok=True)

        stem = f"{uid}_{safe_gene}"

        # ── figures (PNG + PDF + SVG + 8 individual panels) ──────────────────
        png_path = os.path.join(folder, f"{stem}_analysis.png")
        self.visualize_protein(result, output_file=png_path)

        # ── Excel workbook ────────────────────────────────────────────────────
        xlsx_path = os.path.join(folder, f"{stem}_report.xlsx")
        self.export_excel(result, xlsx_path)

        # ── Interactive HTML ──────────────────────────────────────────────────
        html_path = os.path.join(folder, f"{stem}_report.html")
        self.export_html(result, html_path)

        # ── Structured JSON ───────────────────────────────────────────────────
        json_path = os.path.join(folder, f"{stem}_data.json")
        self.export_json_structured(result, json_path)

        # ── TSV site table ────────────────────────────────────────────────────
        tsv_path = os.path.join(folder, f"{stem}_sites.tsv")
        self.export_tsv(result, tsv_path)

        print(f"\n  Results folder → {folder}/")
        for fname in sorted(os.listdir(folder)):
            if not fname.startswith("."):
                size = os.path.getsize(os.path.join(folder, fname))
                if os.path.isdir(os.path.join(folder, fname)):
                    n = len(os.listdir(os.path.join(folder, fname)))
                    print(f"     ├── {fname}/  ({n} files)")
                else:
                    print(f"     ├── {fname}  ({size//1024} KB)")
        return folder

    # ── Enrich result with external data ─────────────────────────────────────

    def enrich_result(self, result: Dict, skip: List[str] = None) -> Dict:
        """
        Add disorder, known PTMs, pathways, and literature to a result.
        Returns enriched result dict.

        Parameters
        ----------
        result : dict  – output from analyze()
        skip   : list  – modules to skip, e.g. ['kegg', 'blast']
        """
        from .enrichment import EnrichmentEngine
        if not hasattr(self, "_enrichment_engine"):
            self._enrichment_engine = EnrichmentEngine()
        print(f"\n  Enriching: {result.get('protein_name','')}")
        return self._enrichment_engine.enrich(result, skip=skip)

    # ── Structured JSON export ────────────────────────────────────────────────

    def export_json_structured(self, result: Dict,
                                output_file: str = "parbm_data.json") -> str:
        """Rich JSON export matching the full tool specification schema."""
        import os
        from datetime import datetime

        enrich   = result.get("enrichment", {})
        dis_data = enrich.get("disorder", {})
        dis_sc   = dis_data.get("scores", [])

        predictions = []
        for site in result.get("paryation_sites", []):
            dis_val = (dis_sc[site.position - 1]
                       if dis_sc and site.position <= len(dis_sc) else None)
            predictions.append({
                "position":            site.position,
                "residue":             site.residue,
                "motif_type":          site.motif_type,
                "confidence":          round(site.score, 4),
                "tier":                ("high" if site.score >= 0.7
                                        else "medium" if site.score >= 0.5
                                        else "low"),
                "disorder_score":      round(dis_val, 3) if dis_val is not None else None,
                "in_disordered_region": (dis_val > 0.5 if dis_val is not None else None),
            })

        known_ptm = [
            {
                "position":     s.get("position"),
                "residue":      s.get("residue"),
                "modification": s.get("modification"),
                "source":       "PHOSPHO.ELM",
            }
            for s in enrich.get("known_ptm_sites", [])
        ]

        pathways = []
        for src, pws in enrich.get("pathways", {}).items():
            for pw in pws:
                pathways.append({**pw, "source": src})

        literature = [
            {
                "title":    pub.get("title"),
                "authors":  pub.get("authors"),
                "journal":  pub.get("journal"),
                "year":     pub.get("year"),
                "pmid":     pub.get("pmid"),
                "doi":      pub.get("doi"),
                "cited_by": pub.get("cited_by"),
            }
            for pub in enrich.get("literature", [])
        ]

        data = {
            "schema_version": "3.0",
            "generated":      datetime.now().isoformat(),
            "protein": {
                "uniprot_id":    result.get("uniprot_id"),
                "protein_name":  result.get("protein_name"),
                "gene_name":     result.get("gene_name"),
                "organism":      result.get("organism"),
                "sequence_length": result.get("sequence_length"),
                "sequence":      result.get("sequence"),
                "input_type":    result.get("input_type"),
            },
            "summary": result.get("summary", {}),
            "predictions":  predictions,
            "domains":      [asdict(d) for d in result.get("all_domains", [])],
            "parbm_domains": [asdict(d) for d in result.get("parbm_domains", [])],
            "enrichment": {
                "disorder": {
                    "method":              dis_data.get("method"),
                    "mean_disorder":       dis_data.get("mean_disorder"),
                    "disordered_fraction": dis_data.get("disordered_fraction"),
                },
                "known_ptm_sites": known_ptm,
                "pathways":        pathways,
                "literature":      literature,
            },
        }

        with open(output_file, "w") as f:
            json.dump(data, f, indent=2)
        print(f"  ✓ JSON exported  → {output_file}")
        return output_file

    # ── TSV site export ───────────────────────────────────────────────────────

    def export_tsv(self, result: Dict,
                   output_file: str = "parbm_sites.tsv") -> str:
        """Tab-separated site table with disorder and known-PTM columns."""
        import csv

        enrich   = result.get("enrichment", {})
        dis_sc   = enrich.get("disorder", {}).get("scores", [])
        known_pos = {s["position"] for s in enrich.get("known_ptm_sites", [])
                     if s.get("position")}

        header = [
            "Position", "Residue", "Motif_Type", "Confidence_Score", "Tier",
            "Disorder_Score", "In_Disordered_Region",
            "Experimental_Evidence", "Source",
            "Protein_Name", "UniProt_ID", "Sequence_Length",
        ]

        with open(output_file, "w", newline="") as f:
            w = csv.writer(f, delimiter="\t")
            w.writerow(header)
            for site in result.get("paryation_sites", []):
                dis_val  = (dis_sc[site.position - 1]
                            if dis_sc and site.position <= len(dis_sc) else "")
                dis_bool = (str(dis_val > 0.5) if dis_val != "" else "")
                tier     = ("high" if site.score >= 0.7
                            else "medium" if site.score >= 0.5 else "low")
                exp_ev   = site.position in known_pos
                w.writerow([
                    site.position, site.residue, site.motif_type,
                    f"{site.score:.4f}", tier,
                    f"{dis_val:.3f}" if dis_val != "" else "",
                    dis_bool,
                    "TRUE" if exp_ev else "FALSE",
                    "PHOSPHO.ELM" if exp_ev else "",
                    result.get("protein_name", ""),
                    result.get("uniprot_id") or "",
                    result.get("sequence_length", ""),
                ])

        print(f"  ✓ TSV exported   → {output_file}")
        return output_file

    # ── HTML report export ────────────────────────────────────────────────────

    def export_html(self, result: Dict,
                    output_file: str = "parbm_report.html") -> Optional[str]:
        """Generate interactive HTML report."""
        from .html_report import HTMLReportGenerator
        return HTMLReportGenerator().generate(result, output_file)
